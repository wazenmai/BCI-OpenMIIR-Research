import os
import mne
import openpyxl
import numpy as np
import matplotlib.pyplot as plt

from asrpy.asrpy import ASR
# from mne_icalabel import label_components

dir_path = "./openmiir/eeg/mne/"
DEFAULT_VERSION = 1
CONDITIONS = ['cued', 'non-cued', 'free']
STIMULUS_IDS = [1, 2, 3, 4, 11, 12, 13, 14, 21, 22, 23, 24]

def compute_ica(beat_epochs, random_seed=42):
    random_state = np.random.RandomState(random_seed)
    ica = mne.preprocessing.ICA(n_components=0.99, method='infomax', fit_params=dict(extended=True), random_state=random_state)
    ica.fit(beat_epochs)
    return ica

## aux functions to be moved to lib
def plot_ica_components(ica, picks=None, topomap_size=3.5):
    if picks is None:
        n_components = ica.mixing_matrix_.shape[1]
        picks = list(range(n_components))
    if len(picks) == 0:
        print('nothing selected for plotting')
        return
    ica.plot_components(picks=picks, ch_type='eeg', title='', colorbar=True, show=False)
    axes = plt.gcf()
    axes.set_size_inches(min(len(picks), 5) * topomap_size, max(len(picks)/5.0, 1) * topomap_size)
    plt.show()

def merge_trial_and_audio_onsets(raw, use_audio_onsets=True, inplace=True, stim_channel='STI 014', verbose=None):
    events = mne.find_events(raw, stim_channel='STI 014', shortest_event=0)

    merged = list()
    last_trial_event = None
    for i, event in enumerate(events):
        etype = event[2]
        if etype < 1000 or etype == 1111: # trial or noise onset
            if use_audio_onsets and events[i+1][2] == 1000: # followed by audio onset
                onset = events[i+1][0]
                merged.append([onset, 0, etype])
                if verbose:
                    print('merged {} + {} = {}'.format(event, events[i+1], merged[-1]))
            else:
                # either we are not interested in audio onsets or there is none
                merged.append(event)
                if verbose:
                    print('kept {}'.format(merged[-1]))
        # audio onsets (etype == 1000) are not copied
        if etype > 1111: # other events (keystrokes)
            merged.append(event)
            if verbose:
                print('kept other {}'.format(merged[-1]))

    merged = np.asarray(merged, dtype=int)

    if inplace:
        stim_id = raw.ch_names.index(stim_channel)
        raw._data[stim_id,:].fill(0)     # delete data in stim channel
        raw.add_events(merged)

    return merged

def load_stimuli_metadata(data_root=None, version=None, verbose=None):

    if version is None:
        version = DEFAULT_VERSION

    # if data_root is None:
        # data_root = os.path.join(deepthought.DATA_PATH, 'OpenMIIR')

    xlsx_filepath = os.path.join(data_root, 'meta', 'Stimuli_Meta.v{}.xlsx'.format(version))
    book = openpyxl.load_workbook(xlsx_filepath, data_only=True)
    sheet = book.worksheets[0]

    if verbose:
        print('Loading stimulus metadata from {}'.format(xlsx_filepath))

    meta = dict()
    for i in range(2, 14):
        stimulus_id = int(sheet.cell(i, 1).value)
        meta[stimulus_id] = {
            'id' : stimulus_id,
            'label' : sheet.cell(i,2).value,
            'audio_file' : sheet.cell(i,3).value,
            'cue_file' : sheet.cell(i,3).value.replace('.wav', '_cue.wav'),
            'length_with_cue' : sheet.cell(i,4).value,
            'length_of_cue' : sheet.cell(i,5).value,
            'length_without_cue' : sheet.cell(i,6).value,
            'length_of_cue_only' : sheet.cell(i,7).value,
            'cue_bpm' : int(sheet.cell(i,8).value),
            'beats_per_bar' : int(sheet.cell(i,9).value),
            'num_bars' : int(sheet.cell(i,15).value),
            'cue_bars' : int(sheet.cell(i,16).value),
            'bpm' : int(sheet.cell(i,17).value),
            'approx_bar_length' : sheet.cell(i,12).value,
        }

        if version == 2:
            meta[stimulus_id]['bpm'] = meta[stimulus_id]['cue_bpm'] # use cue bpm

    return meta

def load_beat_times(stimulus_id, cue=False, data_root=None, verbose=None, version=None):

    if version is None:
        version = DEFAULT_VERSION

    if data_root is None:
        print("error")
        return

    if cue:
        beats_filepath = os.path.join(data_root, 'meta',
                                      'beats.v{}'.format(version),
                                      '{}_cue_beats.txt'.format(stimulus_id))
    else:
        beats_filepath = os.path.join(data_root, 'meta',
                                      'beats.v{}'.format(version),
                                      '{}_beats.txt'.format(stimulus_id))

    with open(beats_filepath, 'r') as f:
        lines = f.readlines()

    beats = []
    for line in lines:
        if not line.strip().startswith('#'):
            beats.append(float(line.strip()))
    beats = np.asarray(beats)

    if verbose:
        print('Read {} beat times from {}'.format(len(beats), beats_filepath))

    return beats

def load_stimuli_metadata_map(key=None, data_root=None, verbose=None, version=None):

    if version is None:
        version = DEFAULT_VERSION

    # handle special case for beats
    if key == 'cue_beats':
        key = 'beats'
        cue = True
    else:
        cue = False

    if key == 'beats':
        map = dict()
        for stimulus_id in STIMULUS_IDS:
            map[stimulus_id] = load_beat_times(stimulus_id,
                                               cue=cue,
                                               data_root=data_root,
                                               verbose=None,
                                               version=version)
        return map

    meta = load_stimuli_metadata(data_root, version=version)

    if key is None:
        return meta  # return everything

    map = dict()
    for stimulus_id in STIMULUS_IDS:
        map[stimulus_id] = meta[stimulus_id][key]

    return map

def beat_event_id_generator(stimulus_id, condition, cue, beat_count):
    if cue:
        cue = 0
    else:
        cue = 10
    return 100000 + stimulus_id * 1000 + condition * 100 + cue + beat_count
        
def decode_event_id(event_id):
    if event_id < 1000:
        stimulus_id = int(event_id / 10)
        condition = int(event_id % 10)
        return stimulus_id, condition
    else:
        return event_id
    
def generate_beat_events(raw, trial_events):
    meta = load_stimuli_metadata(data_root="./openmiir")
    beats = load_stimuli_metadata_map(data_root="./openmiir", key='beats')
    cue_beats = load_stimuli_metadata_map(data_root="./openmiir", key='cue_beats')

    ## determine the number of cue beats
    num_cue_beats = dict()
    for stimulus_id in STIMULUS_IDS:
        num_cue_beats[stimulus_id] = \
            meta[stimulus_id]['beats_per_bar'] * meta[stimulus_id]['cue_bars']
    # print(num_cue_beats)
    
    beat_events = []
    
    ## helper function to add a single beat event
    def add_beat_event(etime, stimulus_id, condition, beat_count, cue=False):
        etype = beat_event_id_generator(stimulus_id, condition, cue, beat_count)
        beat_events.append([etime, 0, etype])
        # print(beat_events[-1])

    ## helper function to add a batch of beat events
    def add_beat_events(etimes, stimulus_id, condition, cue=False):
        beats_per_bar = meta[stimulus_id]['beats_per_bar']
        for i, etime in enumerate(etimes):
            beat_count = (i % beats_per_bar) + 1
            add_beat_event(etime, stimulus_id, condition, beat_count, cue)

    use_audio_onset = True
    include_cue_beats = True
    sr = raw.info['sfreq']
    for i, event in enumerate(trial_events):
        etype = event[2]
        etime = event[0]

        # print('{:4d} at {:8d}'.format(etype, etime))

        if etype >= 1000: # stimulus_id + condition
            continue

        stimulus_id, condition = decode_event_id(etype)

        trial_start = etime # default: use trial onset
        if use_audio_onset and condition < 3:
            # Note: conditions 3 and 4 have no audio cues
            next_event = trial_events[i+1]
            if next_event[2] == 1000: # only use if audio onset
                trial_start = next_event[0]

        # print('Trial start at {}'.format(trial_start))
        # print(stimulus_id, condition)
        if condition < 3: # cued
            offset = sr * meta[stimulus_id]['length_of_cue']

            if include_cue_beats:
                cue_beat_times = trial_start + np.floor(sr * cue_beats[stimulus_id])
                cue_beat_times = cue_beat_times[:num_cue_beats[stimulus_id]]  # truncate at num_cue_beats
                cue_beat_times = np.asarray(cue_beat_times, dtype=int)
                # print(cue_beat_times)
                add_beat_events(cue_beat_times, stimulus_id, condition, cue=True)
        else:
            offset = 0 # no cue

        beat_times = trial_start + offset + np.floor(sr * beats[stimulus_id])
        beat_times = np.asarray(beat_times, dtype=int)
        # print(beat_times[:5], '...')
        add_beat_events(beat_times, stimulus_id, condition)

    beat_events = np.asarray(beat_events, dtype=int)
    return beat_events

## aux function to score EEG channels by EOG correlation
def find_eog_artifact_sources(ica, raw, plot=True, verbose=None):

    eog_picks = mne.pick_types(raw.info, meg=False, eeg=False, eog=True, stim=False)
    eog_inds_set = set()
    multi_scores = list()
    for ch in eog_picks:
        ch_name = raw.ch_names[ch]
        eog_inds, scores = ica.find_bads_eog(raw, str(ch_name), verbose=verbose)
    #     print eog_inds, scores
        if plot:
            ica.plot_scores(scores, exclude=eog_inds, title='EOG artifact sources (red) for channel {}'.format(ch_name))

        multi_scores.append(scores)
        eog_inds_set.update(eog_inds)
    multi_scores = np.vstack(multi_scores)
    # print multi_scores.shape

    # IMPORTANT: due to a + operation meant to concatenate lists, ica.excluded and eog_inds must be lists, not ndarrays
    # see _pick_sources() in ica.py, line 1160
    eog_inds = list(eog_inds_set)
    scores = np.max(np.abs(multi_scores), axis=0).squeeze()

    print('suggested EOG artifact channels: ', eog_inds)
    print('EOG artifact component scores: ', scores[eog_inds])
    return eog_inds, scores
    
    # self.merge_artifact_components() # update 
    
## aux function for readable one-liner code in notebook
def merge_artifact_components(eog_exclude_inds=None, audo_exclude_inds=None):

    sets = list()
    if eog_exclude_inds:
        sets.append(eog_exclude_inds)
    if auto_exclude_inds:
        sets.append(auto_exclude_inds)

    if len(sets) == 1:
        merged = sets[0]
    else:
        print('merging', sets)
        merged = set()
        for s in sets:
            for e in s:
                merged.add(e)
        merged = sorted(list(merged))
    return merged

    # self.suggested_artifact_components = merged
    
def auto_detect_artifact_components(ica, beat_epochs, eog_exclude_inds):

    data = beat_epochs

    """
    data: raw, epochs or evoked
    """

    exclude_old = ica.exclude  # store old setting
    ica.exclude = []
    ica.detect_artifacts(data)
    auto_exclude = ica.exclude
    ica.exclude = exclude_old  # restore old setting

    suggested_artifact_components = merge_artifact_components(eog_exclude_inds=eog_exclude_inds, auto_exclude_indx=auto_exclude) # update combination
    return suggested_artifact_components


def read_data(dir_path, filepath, filter=False, reconstructed=False):
    raw = mne.io.read_raw_fif(dir_path + filepath, preload=True)
    subject = filepath.split('-')[0]
    if not reconstructed:
        if subject in ['P09', 'P11', 'P12', 'P13', 'P14']:
            raw.set_eeg_reference(['EXG5', 'EXG6'])
            raw.drop_channels(['EXG5', 'EXG6'])
        else:
            raw.set_eeg_reference()
    else:
        raw.set_eeg_reference()
    raw.set_montage('biosemi64', on_missing='ignore')
    if len(raw.info['bads']) > 0:
        raw.interpolate_bads()

    if filter:
        # bandpass-filter
        eeg_picks = mne.pick_types(raw.info, meg=False, eeg=True, eog=False, stim=False, exclude=[])
        fmin = 1
        fmax = 30
        raw.filter(fmin, fmax, picks=eeg_picks, filter_length='10s',
                  l_trans_bandwidth=0.1, h_trans_bandwidth=0.5, method='iir', verbose=False)

    # merge
    merge_trial_and_audio_onsets(raw, use_audio_onsets=True, inplace=True)

    # recompute trial_events and times
    trial_events = mne.find_events(raw, stim_channel='STI 014', shortest_event=0)
    trial_event_times = raw.times[trial_events[:,0]]
    return raw, trial_events

def generate_data_and_label(raw, trial_events, condition, psd=False, epoch=False):
    """
    Parameter
    - raw
    - events (trial events)
    - condition
      1. Perception vs. Imagination
      2. Pop song vs. Instrument music (perception, without lyrics)
      3. Lyrics vs. No Lyrics vs. Instrument
      4. Cued vs. Not Cued
      5. Stimulus classification
    Return
    - data array, label array, group array (for CV)
    """
    picks = mne.pick_types(raw.info, meg=False, eeg=True, eog=False, stim=False, exclude=[])
    data_array, label_array, group_array = None, None, None

    # For beat event
    tmin = -0.2  # start of each epoch (200ms before the trigger)
    tmax = 0.8  # end of each epoch (600ms after the trigger) - longest beat is 0.57s long
    
    # For listening event
    tmin = -1
    tmax = 5

    data_epochs = []

    if condition == 1:
        tmin = -1
        tmax = 5
        listen_with_cue_epochs = mne.Epochs(raw, trial_events, [id*10 + 1 for id in STIMULUS_IDS],
                          tmin, tmax, preload=True,
                          proj=False, picks=picks, verbose=False)
        data_epochs.append(listen_with_cue_epochs)
        listen_data = listen_with_cue_epochs.get_data()
        if psd:
            frequency = listen_with_cue_epochs.compute_psd()
            listen_data = np.concatenate((listen_data, frequency), axis=2)
        imagine_with_cue_epochs = mne.Epochs(raw, trial_events, [id*10 + 2 for id in STIMULUS_IDS],
                          tmin, tmax, preload=True,
                          proj=False, picks=picks, verbose=False)
        data_epochs.append(imagine_with_cue_epochs)
        imagine_data = imagine_with_cue_epochs.get_data()
        if psd:
            frequency = imagine_with_cue_epochs.compute_psd()
            imagine_data = np.concatenate((imagine_data, frequency), axis=2)
        listen_epoch_labels = [0] * listen_data.shape[0]
        imagine_epoch_labels = [1] * imagine_data.shape[0]

        # concat data and label
        data_list = np.concatenate((listen_data, imagine_data), axis=0)
        label_list = listen_epoch_labels + imagine_epoch_labels

        # make group
        group_list = [[i] * 5 for i in range(12)]
        group_list = group_list + group_list

        group_array = np.hstack(group_list)
        data_array = data_list
        label_array = np.array(label_list)

    # Fast tempo vs. Slow Tempo (Perception + Imagination)
    elif condition == 2: 
        tmin = -1
        tmax = 5
        fast_id = [1, 2, 3, 11, 12, 13]
        condition_id = [1, 2]
        fast_event_id = []
        for stimulus in fast_id:
            for condition in condition_id:
                fast_event_id.append(stimulus*10 + condition)
        fast_epochs = mne.Epochs(raw, trial_events, fast_event_id,
                        tmin, tmax, preload=True,
                        proj=False, picks=picks, verbose=False)
        data_epochs.append(fast_epochs)
        fast_data = fast_epochs.get_data()
        if psd:
            frequency = fast_epochs.compute_psd()
            fast_data = np.concatenate((fast_data, frequency), axis=2)
        slow_id = [4, 14, 21, 22, 23, 24]
        slow_event_id = []
        for stimulus in slow_id:
            for condition in condition_id:
                slow_event_id.append(stimulus*10 + condition)
        slow_epochs = mne.Epochs(raw, trial_events, slow_event_id,
                        tmin, tmax, preload=True,
                        proj=False, picks=picks, verbose=False)
        data_epochs.append(slow_epochs)
        slow_data = slow_epochs.get_data()
        if psd:
            frequency = slow_epochs.compute_psd()
            slow_data = np.concatenate((fast_data, frequency), axis=2)
        
        fast_epoch_labels = [0] * fast_epochs.shape[0]
        slow_epoch_labels = [1] * slow_epochs.shape[0]

        # concat data and label
        data_list = np.concatenate((fast_data, slow_data), axis=0)
        label_list = fast_epoch_labels + slow_epoch_labels

        # make group
        group_list = [[i] * 5 for i in range(12)]
        group_list = group_list + group_list

        group_array = np.hstack(group_list)
        data_array = data_list
        label_array = np.array(label_list)

    # Song identification (all, perception, imagination)
    elif condition == 3 or condition == 4 or condition == 5:
        tmin = -1
        tmax = 5
        data_list = []
        for stimulus in STIMULUS_IDS:
            if condition == 3:
                event_ids = [stimulus*10 + 1, stimulus*10 + 2]
            elif condition == 4:
                event_ids = [stimulus*10 + 1]
            elif condition == 5:
                event_ids = [stimulus*10 + 2]
            epochs = mne.Epochs(raw, trial_events, event_ids,
                            tmin, tmax, preload=True,
                            proj=False, picks=picks, verbose=False)
            data_epochs.append(epochs)
            data = epochs.get_data()
            if psd:
                frequency = epochs.compute_psd()
                data = np.concatenate((data, frequency), axis=2)
            data_list.append(data)
        labels = []
        for i, stimulus in enumerate(STIMULUS_IDS):
            if condition == 3:
                labels.extend([i] * 10)
            elif condition == 4 or condition == 5:
                labels.extend([i] * 5)
        data = np.concatenate(data_list, axis=0)

        # make group
        if condition == 3:
            group_list = [[i] * 10 for i in range(12)]
        elif condition == 4 or condition == 5:
            group_list = [[i] * 5 for i in range(12)]

        group_array = np.hstack(group_list)
        data_array = data
        label_array = np.array(labels)


    print("data: {}, label: {}, group: {}".format(data_array.shape, label_array.shape, group_array.shape))

    if epoch:
        return data_array, label_array, group_array, data_epochs
    return data_array, label_array, group_array

def tight_layout(pad=1.2, h_pad=None, w_pad=None, fig=None):
    """ Adjust subplot parameters to give specified padding.

    Note. For plotting please use this function instead of plt.tight_layout

    Parameters
    ----------
    pad : float
        padding between the figure edge and the edges of subplots, as a
        fraction of the font-size.
    h_pad : float
        Padding height between edges of adjacent subplots.
        Defaults to `pad_inches`.
    w_pad : float
        Padding width between edges of adjacent subplots.
        Defaults to `pad_inches`.
    fig : instance of Figure
        Figure to apply changes to.
    """
    import matplotlib.pyplot as plt
    if fig is None:
        fig = plt.gcf()

    try:  # see https://github.com/matplotlib/matplotlib/issues/2654
        fig.canvas.draw()
        fig.tight_layout(pad=pad, h_pad=h_pad, w_pad=w_pad)
    except:
        msg = ('Matplotlib function \'tight_layout\'%s.'
               ' Skipping subpplot adjusment.')
        if not hasattr(plt, 'tight_layout'):
            case = ' is not available'
        else:
            case = (' is not supported by your backend: `%s`'
                    % plt.get_backend())
        warn(msg % case)

def plot_ica_overlay_evoked(evoked, evoked_cln, title, show):
    """
    workaround for https://github.com/mne-tools/mne-python/issues/1819
    copied from mne.viz.ica._plot_ica_overlay_evoked()

    Plot evoked after and before ICA cleaning

    Parameters
    ----------
    ica : instance of mne.preprocessing.ICA
        The ICA object.
    epochs : instance of mne.Epochs
        The Epochs to be regarded.
    show : bool
        If True, all open plots will be shown.

    Returns
    -------
    fig : instance of pyplot.Figure
    """
    ch_types_used = [c for c in ['mag', 'grad', 'eeg'] if c in evoked]
    n_rows = len(ch_types_used)
    ch_types_used_cln = [c for c in ['mag', 'grad', 'eeg'] if
                         c in evoked_cln]

    if len(ch_types_used) != len(ch_types_used_cln):
        raise ValueError('Raw and clean evokeds must match. '
                         'Found different channels.')

    fig, axes = plt.subplots(n_rows, 1)
    fig.suptitle('Average signal before (red) and after (black) ICA')
    axes = axes.flatten() if isinstance(axes, np.ndarray) else axes

    evoked.plot(axes=axes, show=False)

    for ax in fig.axes:
        [l.set_color('r') for l in ax.get_lines()]

    fig.canvas.draw()
    evoked_cln.plot(axes=axes, show=show)
    tight_layout(fig=fig)

    if show:
        plt.show()

    fig.subplots_adjust(top=0.90)
    fig.canvas.draw()

## Assess component selection and unmixing quality
def assess_unmixing_quality(ica, beat_epochs, eog_epochs, raw, verbose=None): # eog_evoked=None, raw=None, evoked=None, verbose=None):
    eog_evoked = eog_epochs.average()
    data = beat_epochs

    # if isinstance(data, mne.epochs._BaseEpochs):
    #     evoked = data.average()
    # else:
    #     evoked = None
    evoked = data.average()

    if eog_evoked is not None:
        print('Assess impact on average EOG artifact:')
        ica.plot_sources(eog_evoked)  # plot EOG sources + selection

        print('Assess cleaning of EOG epochs:')

        # Note: this method appears to be broken! Lines that should be red are drawn in black
        # ica.plot_overlay(eog_evoked, exclude=ica.exclude)

        # workaroud
        evoked_cln = ica.apply(eog_evoked, exclude=ica.exclude).copy()
        plot_ica_overlay_evoked(evoked=eog_evoked, evoked_cln=evoked_cln, title='', show=True)

    if raw is not None:
        print ('Assess impact on raw. Check the amplitudes do not change:')
        ica.plot_overlay(raw)  # EOG artifacts remain

    if evoked is not None:
        print ('Assess impact on evoked. Check the amplitudes do not change:')
        evoked_cln = ica.apply(evoked, exclude=ica.exclude).copy()
        plot_ica_overlay_evoked(evoked=evoked, evoked_cln=evoked_cln, title='', show=True)
  
def remove_artifact(filepath, outputdir):
    raw, events = read_data(filepath, filter=True)
    
    # beat epoching
    trial_events = events
    beat_events = generate_beat_events(raw, trial_events)
    picks = mne.pick_types(raw.info, meg=False, eeg=True, eog=True, stim=True, exclude=[])
    event_id = None # any
    tmin = -0.2  # start of each epoch (200ms before the trigger)
    tmax = 0.8  # end of each epoch (600ms after the trigger) - longest beat is 0.57s long
    detrend = 0 # remove dc
    beat_epochs = mne.Epochs(raw, beat_events, event_id,
                                  tmin, tmax, preload=True,
                                  proj=False, picks=picks, verbose=False)
    
    # EOG epoching
    eog_event_id = 5000
    eog_events = mne.preprocessing.find_eog_events(raw, eog_event_id)

    # create epochs around EOG events
    picks = mne.pick_types(raw.info, meg=False, eeg=True, eog=True, stim=True, exclude=[]) # FIXME
    tmin = -.5
    tmax = .5
    eog_epochs = mne.Epochs(raw, events=eog_events, event_id=eog_event_id,
                        tmin=tmin, tmax=tmax, proj=False, picks=picks,
                        preload=True, verbose=False)
    
    # compute ica
    ica = compute_ica(beat_epochs)
    
    # find eog artifact
    eog_inds, scores = find_eog_artifact_sources(ica, raw, plot=False)
    eog_idx, eog_score = ica.find_bads_eog(beat_epochs)
    exclude_idx = np.unique(np.hstack((eog_idx, eog_inds)))
    ica.exclude = exclude_idx
    
    # Reconstruct artifact
    ica.apply(raw, exclude=ica.exclude)
    if len(raw.info['bads']) > 0:
        raw.interpolate_bads_eeg() # interpolate bad channels afterwards as they are not processed by the ICA

    # save file
    subject = filepath.split('-')[0]
    raw.save(os.path.join(outputdir, subject+"-rec.fif"))
    ica.save(os.path.join(outputdir, subject+"-ica.fif"))
    
    return raw, beat_epochs, eog_epochs, ica

if __name__ == "__main__":
    output_dir = "/home/wazenmai/Experiment/bci/openmiir/eeg/processed/"

    subjects = []
    for root, dir, files in os.walk(dir_path):
        for f in files:
            if ".fif" in f:
                subjects.append(f)

    for subject in subjects:
        print("Processing subject: {}".format(subject))
        raw, beat_epochs, eog_epochs, ica = remove_artifact(subject, output_dir)
        print("Done processing subject: {}".format(subject))