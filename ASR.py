import os
import mne
import openpyxl
import numpy as np
import matplotlib.pyplot as plt
import time


from asrpy.asrpy import ASR

dir_path = "./openmiir/eeg/mne/"
raw = mne.io.read_raw_fif(dir_path + "P01-raw.fif", preload=True)
raw.set_eeg_reference()
print(raw)
print(raw.info)

# interpolate bad channels
raw.interpolate_bads(origin=(0.0, 0.0, 0.0))

# bandpass-filter
eeg_picks = mne.pick_types(raw.info, meg=False, eeg=True, eog=False, stim=False, exclude=[])
fmin = 1
fmax = 30
raw.filter(fmin, fmax, picks=eeg_picks, filter_length='10s',
           l_trans_bandwidth=0.1, h_trans_bandwidth=0.5, method='iir',
           verbose=True)
print("hey??")
DEFAULT_VERSION = 1
CONDITIONS = ['cued', 'non-cued', 'free']
STIMULUS_IDS = [1, 2, 3, 4, 11, 12, 13, 14, 21, 22, 23, 24]

def load_stimuli_metadata(data_root=None, version=None, verbose=None):

    if version is None:
        version = DEFAULT_VERSION

    if data_root is None:
        data_root = os.path.join(deepthought.DATA_PATH, 'OpenMIIR')

    xlsx_filepath = os.path.join(data_root, 'meta', 'Stimuli_Meta.v{}.xlsx'.format(version))
    book = openpyxl.load_workbook(xlsx_filepath, data_only=True)
    sheet = book.worksheets[0]

    if verbose:
        log.info('Loading stimulus metadata from {}'.format(xlsx_filepath))

    meta = dict()
    for i in range(2, 14):
        stimulus_id = int(sheet.cell(i, 1).value)
        meta[stimulus_id] = {
            'id' : stimulus_id,
            'label' : sheet.cell(i,2).value,
            'audio_file' : sheet.cell(i,3).value,
            'cue_file' : sheet.cell(i,3).value.replace('.wav', '_cue.wav'),
            'length_with_cue' : sheet.cell(i,4).value,
            'length_of_cue' : sheet.cell(i,5).value,
            'length_without_cue' : sheet.cell(i,6).value,
            'length_of_cue_only' : sheet.cell(i,7).value,
            'cue_bpm' : int(sheet.cell(i,8).value),
            'beats_per_bar' : int(sheet.cell(i,9).value),
            'num_bars' : int(sheet.cell(i,15).value),
            'cue_bars' : int(sheet.cell(i,16).value),
            'bpm' : int(sheet.cell(i,17).value),
            'approx_bar_length' : sheet.cell(i,12).value,
        }

        if version == 2:
            meta[stimulus_id]['bpm'] = meta[stimulus_id]['cue_bpm'] # use cue bpm

    return meta

def load_beat_times(stimulus_id, cue=False, data_root=None, verbose=None, version=None):

    if version is None:
        version = DEFAULT_VERSION

    if data_root is None:
        data_root = os.path.join(deepthought.DATA_PATH, 'OpenMIIR')

    if cue:
        beats_filepath = os.path.join(data_root, 'meta',
                                      'beats.v{}'.format(version),
                                      '{}_cue_beats.txt'.format(stimulus_id))
    else:
        beats_filepath = os.path.join(data_root, 'meta',
                                      'beats.v{}'.format(version),
                                      '{}_beats.txt'.format(stimulus_id))

    with open(beats_filepath, 'r') as f:
        lines = f.readlines()

    beats = []
    for line in lines:
        if not line.strip().startswith('#'):
            beats.append(float(line.strip()))
    beats = np.asarray(beats)

    if verbose:
        print('Read {} beat times from {}'.format(len(beats), beats_filepath))

    return beats

def load_stimuli_metadata_map(key=None, data_root=None, verbose=None, version=None):

    if version is None:
        version = DEFAULT_VERSION

    # handle special case for beats
    if key == 'cue_beats':
        key = 'beats'
        cue = True
    else:
        cue = False

    if key == 'beats':
        map = dict()
        for stimulus_id in STIMULUS_IDS:
            map[stimulus_id] = load_beat_times(stimulus_id,
                                               cue=cue,
                                               data_root=data_root,
                                               verbose=None,
                                               version=version)
        return map

    meta = load_stimuli_metadata(data_root, version=version)

    if key is None:
        return meta  # return everything

    map = dict()
    for stimulus_id in STIMULUS_IDS:
        map[stimulus_id] = meta[stimulus_id][key]

    return map

def beat_event_id_generator(stimulus_id, condition, cue, beat_count):
    if cue:
        cue = 0
    else:
        cue = 10
    return 100000 + stimulus_id * 1000 + condition * 100 + cue + beat_count
       
def decode_event_id(event_id):
    if event_id < 1000:
        stimulus_id = int(event_id / 10)
        condition = int(event_id % 10)
        return stimulus_id, condition
    else:
        return event_id
    
def generate_beat_events(raw, trial_events):
    meta = load_stimuli_metadata(data_root="./openmiir")
    beats = load_stimuli_metadata_map(data_root="./openmiir", key='beats')
    cue_beats = load_stimuli_metadata_map(data_root="./openmiir", key='cue_beats')

    ## determine the number of cue beats
    num_cue_beats = dict()
    for stimulus_id in STIMULUS_IDS:
        num_cue_beats[stimulus_id] = \
            meta[stimulus_id]['beats_per_bar'] * meta[stimulus_id]['cue_bars']
    print(num_cue_beats)
    
    beat_events = []
    
    ## helper function to add a single beat event
    def add_beat_event(etime, stimulus_id, condition, beat_count, cue=False):
        etype = beat_event_id_generator(stimulus_id, condition, cue, beat_count)
        beat_events.append([etime, 0, etype])
        # print(beat_events[-1])

    ## helper function to add a batch of beat events
    def add_beat_events(etimes, stimulus_id, condition, cue=False):
        beats_per_bar = meta[stimulus_id]['beats_per_bar']
        for i, etime in enumerate(etimes):
            beat_count = (i % beats_per_bar) + 1
            add_beat_event(etime, stimulus_id, condition, beat_count, cue)

    use_audio_onset = True
    include_cue_beats = True
    sr = raw.info['sfreq']
    for i, event in enumerate(trial_events):
        etype = event[2]
        etime = event[0]

        # print('{:4d} at {:8d}'.format(etype, etime))

        if etype >= 1000: # stimulus_id + condition
            continue

        stimulus_id, condition = decode_event_id(etype)

        trial_start = etime # default: use trial onset
        if use_audio_onset and condition < 3:
            # Note: conditions 3 and 4 have no audio cues
            next_event = trial_events[i+1]
            if next_event[2] == 1000: # only use if audio onset
                trial_start = next_event[0]

        # print('Trial start at {}'.format(trial_start))
        print(stimulus_id, condition)
        if condition < 3: # cued
            offset = sr * meta[stimulus_id]['length_of_cue']

            if include_cue_beats:
                cue_beat_times = trial_start + np.floor(sr * cue_beats[stimulus_id])
                cue_beat_times = cue_beat_times[:num_cue_beats[stimulus_id]]  # truncate at num_cue_beats
                cue_beat_times = np.asarray(cue_beat_times, dtype=int)
                # print(cue_beat_times)
                add_beat_events(cue_beat_times, stimulus_id, condition, cue=True)
        else:
            offset = 0 # no cue

        beat_times = trial_start + offset + np.floor(sr * beats[stimulus_id])
        beat_times = np.asarray(beat_times, dtype=int)
        # print(beat_times[:5], '...')
        add_beat_events(beat_times, stimulus_id, condition)

    beat_events = np.asarray(beat_events, dtype=int)
    return beat_events

## aux function to score EEG channels by EOG correlation
def find_eog_artifact_sources(ica, raw, plot=True, verbose=None):

    eog_picks = mne.pick_types(raw.info, meg=False, eeg=False, eog=True, stim=False)
    eog_inds_set = set()
    multi_scores = list()
    for ch in eog_picks:
        ch_name = raw.ch_names[ch]
        eog_inds, scores = ica.find_bads_eog(raw, str(ch_name), verbose=verbose)
    #     print eog_inds, scores
        if plot:
            ica.plot_scores(scores, exclude=eog_inds, title='EOG artifact sources (red) for channel {}'.format(ch_name))

        multi_scores.append(scores)
        eog_inds_set.update(eog_inds)
    multi_scores = np.vstack(multi_scores)
    # print multi_scores.shape

    # IMPORTANT: due to a + operation meant to concatenate lists, ica.excluded and eog_inds must be lists, not ndarrays
    # see _pick_sources() in ica.py, line 1160
    eog_inds = list(eog_inds_set)
    scores = np.max(np.abs(multi_scores), axis=0).squeeze()

    print('suggested EOG artifact channels: ', eog_inds)
    print('EOG artifact component scores: ', scores[eog_inds])
    return eog_inds, scores
    
    # self.merge_artifact_components() # update 
    
## aux function for readable one-liner code in notebook
def merge_artifact_components(eog_exclude_inds=None, audo_exclude_inds=None):

    sets = list()
    if eog_exclude_inds:
        sets.append(eog_exclude_inds)
    if auto_exclude_inds:
        sets.append(auto_exclude_inds)

    if len(sets) == 1:
        merged = sets[0]
    else:
        print('merging', sets)
        merged = set()
        for s in sets:
            for e in s:
                merged.add(e)
        merged = sorted(list(merged))
    return merged

    # self.suggested_artifact_components = merged
    
def auto_detect_artifact_components(ica, beat_epochs, eog_exclude_inds):

    data = beat_epochs

    """
    data: raw, epochs or evoked
    """

    exclude_old = ica.exclude  # store old setting
    ica.exclude = []
    ica.detect_artifacts(data)
    auto_exclude = ica.exclude
    ica.exclude = exclude_old  # restore old setting

    suggested_artifact_components = merge_artifact_components(eog_exclude_inds=eog_exclude_inds, auto_exclude_indx=auto_exclude) # update combination
    return suggested_artifact_components

# trial_events = mne.find_events(raw, stim_channel='STI 014', shortest_event=0)
# beat_events = generate_beat_events(raw, trial_events)
# picks = mne.pick_types(raw.info, meg=False, eeg=True, eog=True, stim=True, exclude=[])
# event_id = None # any
# tmin = -0.2  # start of each epoch (200ms before the trigger)
# tmax = 0.8  # end of each epoch (600ms after the trigger) - longest beat is 0.57s long
# detrend = 0 # remove dc
# beat_epochs = mne.Epochs(raw, beat_events, event_id,
#                               tmin, tmax, preload=True,
#                               proj=False, picks=picks, verbose=False)
# print(beat_epochs)
print("hi")


asr = ASR(sfreq=raw.info["sfreq"], cutoff=20)
asr.fit(raw)
print("asr fitted")
asr_raw = asr.transform(raw)
print("asr transformed")
picks = mne.pick_types(asr_raw.info, meg=False, eeg=True, eog=True, stim=True, exclude=[])
event_id = None # any
tmin = -0.2  # start of each epoch (200ms before the trigger)
tmax = 0.8  # end of each epoch (600ms after the trigger) - longest beat is 0.57s long
detrend = 0 # remove dc
clean_beat_epochs = mne.Epochs(asr_raw, beat_events, event_id,
                              tmin, tmax, preload=True,
                              proj=False, picks=picks, verbose=False)
clean_beat_epochs.save('P01-epo.fif')